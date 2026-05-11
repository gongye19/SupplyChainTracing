#!/usr/bin/env python3
"""
Upload company-level trade data as compact dashboard aggregates.

Source:
  data/pure_company_pair_usd_count.zip (preferred, CSV)
  data/hkust_文件汇总/*.xlsx (legacy fallback)

Targets:
  company_search_stats
  company_monthly_trade_stats
  company_hs_trade_stats
  company_counterparty_trade_stats
  country_origin_trade_stats

Usage:
  python scripts/upload_company_trade_flows.py --clear
"""

from __future__ import annotations

import argparse
import csv
import io
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from sqlalchemy import create_engine, text

from importers.aggregate_writer import write_csvs
from importers.company_pair_csv import build_aggregates as build_company_pair_aggregates
from importers.hkust_company import build_aggregates as build_hkust_aggregates
from importers.schema import TABLE_COLUMNS, create_indexes, create_tables, drop_tables, swap_staging_tables


DATABASE_URL = os.getenv(
    "DATABASE_URL",
    os.getenv(
        "DATABASE_PUBLIC_URL",
        "postgresql://postgres:HcVbYOkDrgXNZPaUpttdATshXWFJWWQe@tramway.proxy.rlwy.net:11626/railway",
    ),
)

DB_CONNECT_ARGS = {
    "connect_timeout": 20,
    "keepalives": 1,
    "keepalives_idle": 30,
    "keepalives_interval": 10,
    "keepalives_count": 3,
    "options": "-c statement_timeout=180000",
}

COPY_BATCH_ROWS = 20_000
EXTRA_TOP_COMPANIES = int(os.getenv("COMPANY_IMPORT_EXTRA_TOP_COMPANIES", "0"))
TOP_COUNTERPARTIES_PER_COMPANY_ROLE = int(os.getenv("COMPANY_IMPORT_TOP_COUNTERPARTIES", "25"))


def copy_csv(engine, table: str, columns: list[str], csv_path: Path) -> int:
    total = 0
    batch_no = 0
    cols = ", ".join(columns)
    copy_sql = f"COPY {table} ({cols}) FROM STDIN WITH (FORMAT CSV)"

    def flush(rows: list[list[str]]) -> None:
        nonlocal total, batch_no
        if not rows:
            return
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerows(rows)
        buffer.seek(0)

        raw = engine.raw_connection()
        try:
            with raw.cursor() as cur:
                cur.copy_expert(copy_sql, buffer)
            raw.commit()
        finally:
            raw.close()

        batch_no += 1
        total += len(rows)
        print(f"  {table} batch {batch_no}: {len(rows):,} rows, total {total:,}", flush=True)

    with csv_path.open("r", newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        batch: list[list[str]] = []
        for row in reader:
            batch.append(row)
            if len(batch) >= COPY_BATCH_ROWS:
                flush(batch)
                batch = []
        flush(batch)

    return total


def validate_loaded_tables(engine, table_suffix: str) -> dict[str, int]:
    tables = {
        "company_search_stats": f"company_search_stats{table_suffix}",
        "company_monthly_trade_stats": f"company_monthly_trade_stats{table_suffix}",
        "company_hs_trade_stats": f"company_hs_trade_stats{table_suffix}",
        "company_counterparty_trade_stats": f"company_counterparty_trade_stats{table_suffix}",
        "country_origin_trade_stats": f"country_origin_trade_stats{table_suffix}",
    }
    with engine.connect() as conn:
        counts = {
            logical_name: conn.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar() or 0
            for logical_name, table_name in tables.items()
        }
        months = conn.execute(
            text(
                f"""
                SELECT MIN(year * 100 + month), MAX(year * 100 + month)
                FROM company_monthly_trade_stats{table_suffix}
                """
            )
        ).fetchone()

    required = ["company_search_stats", "company_monthly_trade_stats", "country_origin_trade_stats"]
    empty_required = [name for name in required if counts[name] <= 0]
    if empty_required:
        raise RuntimeError(f"导入结果异常，关键表为空: {', '.join(empty_required)}")
    if not months or months[0] is None or months[1] is None:
        raise RuntimeError("导入结果异常，无法读取月份范围")

    print("staging 数据库记录数:", counts)
    print(f"staging 月份范围: {months[0]} ~ {months[1]}")
    return counts


def pick_default_source(project_root: Path) -> Path:
    zip_source = project_root / "data" / "pure_company_pair_usd_count.zip"
    if zip_source.exists():
        return zip_source
    return project_root / "data" / "hkust_文件汇总"


def load_brand_company_map(path: Path) -> dict[str, str]:
    import openpyxl

    if not path.exists():
        return {}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["明细"] if "明细" in wb.sheetnames else wb.active
        brand_by_company: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 3 or row[2] is None:
                continue
            brand = "" if row[1] is None else str(row[1]).strip()
            company = str(row[2]).strip()
            if company:
                brand_by_company[company] = brand
        return brand_by_company
    finally:
        wb.close()


def build_source_aggregates(source: Path) -> tuple[dict[str, int], dict[str, dict]]:
    if source.is_file() and source.suffix.lower() in {".zip", ".csv"}:
        return build_company_pair_aggregates(source)
    return build_hkust_aggregates(source)


def main() -> None:
    parser = argparse.ArgumentParser(description="Upload company-level trade data as compact aggregates.")
    parser.add_argument("--source", default=None, help="Source zip/csv file or legacy HKUST XLSX directory")
    parser.add_argument("--data-dir", default=None, help="Directory containing HKUST XLSX files")
    parser.add_argument("--brand-list", default=None, help="Optional Excel brand company list; defaults to 215-brand workbook when present")
    parser.add_argument("--database-url", default=None, help="PostgreSQL connection URL")
    parser.add_argument("--clear", action="store_true", help="Backward compatible no-op; imports always rebuild via staging")
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parents[1]
    source = Path(args.source or args.data_dir).expanduser().resolve() if (args.source or args.data_dir) else pick_default_source(project_root)
    default_brand_list = project_root / "data" / "可匹配品牌企业名录（全量215品牌）.xlsx"
    brand_list = Path(args.brand_list).expanduser().resolve() if args.brand_list else default_brand_list
    database_url = args.database_url or DATABASE_URL

    print("=" * 72)
    print("上传公司级聚合数据到数据库")
    print("=" * 72)
    print("模式: 全量重建")
    print("数据源:", source)
    if brand_list.exists():
        print("品牌白名单:", brand_list)

    if not source.exists():
        print(f"❌ 数据源不存在: {source}")
        sys.exit(1)

    engine = create_engine(database_url, pool_pre_ping=True, connect_args=DB_CONNECT_ARGS)
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        print("✓ 数据库连接成功")
    except Exception as exc:
        print(f"❌ 数据库连接失败: {exc}")
        sys.exit(1)

    start = datetime.now()
    stats, aggregates = build_source_aggregates(source)
    source_brand_companies = set(aggregates.get("brand_companies") or set())
    source_brand_map = dict(aggregates.get("brand_by_company") or {})
    list_brand_map = load_brand_company_map(brand_list)
    brand_by_company = {**source_brand_map, **list_brand_map}
    force_company_names = source_brand_companies | set(list_brand_map)
    if not force_company_names:
        raise RuntimeError("未找到215品牌白名单公司，停止导入，避免误导入空公司看板")
    print("聚合完成:", stats)
    if force_company_names:
        print(
            "215品牌白名单公司:",
            {
                "source_brand_companies": len(source_brand_companies),
                "excel_brand_companies": len(list_brand_map),
                "union": len(force_company_names),
            },
        )
    print(
        "聚合行数:",
        {
            "company_search_stats": "derived",
            "company_monthly_trade_stats": len(aggregates["monthly"]),
            "company_hs_trade_stats": len(aggregates["hs"]),
            "company_counterparty_trade_stats": len(aggregates["counterparty"]),
            "country_origin_trade_stats": len(aggregates["country"]),
        },
    )

    with tempfile.TemporaryDirectory(prefix="company_aggregates_") as tmp:
        paths = write_csvs(
            aggregates,
            Path(tmp),
            top_companies=EXTRA_TOP_COMPANIES,
            top_counterparties_per_company_role=TOP_COUNTERPARTIES_PER_COMPANY_ROLE,
            force_company_names=force_company_names,
            brand_by_company=brand_by_company,
        )
        csv_counts = {table: sum(1 for _ in path.open("r", encoding="utf-8")) for table, path in paths.items()}
        print("CSV 行数:", csv_counts)

        staging_suffix = "__import_" + datetime.now().strftime("%Y%m%d%H%M%S")
        print(f"创建 staging 表 {staging_suffix} ...")
        try:
            create_tables(engine, suffix=staging_suffix)

            for table, path in paths.items():
                copy_csv(engine, f"{table}{staging_suffix}", TABLE_COLUMNS[table], path)

            print("创建 staging 查询索引 ...")
            create_indexes(engine, suffix=staging_suffix)
            validate_loaded_tables(engine, staging_suffix)

            print("切换 staging 表为正式表 ...")
            swap_staging_tables(engine, staging_suffix=staging_suffix)
        except Exception:
            print("导入失败，正式表未切换，正在清理 staging 表 ...")
            drop_tables(engine, suffix=staging_suffix)
            raise

    elapsed = (datetime.now() - start).total_seconds()
    with engine.connect() as conn:
        counts = {
            table: conn.execute(text(f"SELECT COUNT(*) FROM {table}")).scalar()
            for table in [
                "company_search_stats",
                "company_monthly_trade_stats",
                "company_hs_trade_stats",
                "company_counterparty_trade_stats",
                "country_origin_trade_stats",
            ]
        }
        months = conn.execute(
            text("SELECT MIN(year * 100 + month), MAX(year * 100 + month) FROM company_monthly_trade_stats")
        ).fetchone()

    print("\n" + "=" * 72)
    print("上传完成")
    print("=" * 72)
    print("数据库记录数:", counts)
    print(f"月份范围: {months[0]} ~ {months[1]}")
    print(f"耗时: {elapsed:.1f}s")


if __name__ == "__main__":
    main()
