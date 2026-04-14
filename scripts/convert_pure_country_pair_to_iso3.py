#!/usr/bin/env python3
"""
将 pure_country_pair_usd_count 下的进口国/出口国代码统一为 ISO3。

用法:
  python3 scripts/convert_pure_country_pair_to_iso3.py --dry-run
  python3 scripts/convert_pure_country_pair_to_iso3.py --apply
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Optional

import openpyxl
import pycountry


DATA_DIR = Path(__file__).resolve().parents[1] / "data" / "pure_country_pair_usd_count"
CODE_RE = re.compile(r"^(?P<name>.*)\((?P<code>[A-Z]{2,3})\)\s*$")

ISO2_TO_ISO3_MANUAL = {
    "UK": "GBR",
    "EL": "GRC",
}


def to_iso3(code: str) -> Optional[str]:
    code = (code or "").strip().upper()
    if not code:
        return None
    if len(code) == 3:
        return code
    if len(code) == 2:
        if code in ISO2_TO_ISO3_MANUAL:
            return ISO2_TO_ISO3_MANUAL[code]
        country = pycountry.countries.get(alpha_2=code)
        return country.alpha_3 if country else None
    return None


def normalize_country_cell(value: object) -> tuple[object, bool, bool]:
    """
    返回: (new_value, changed, unresolved)
    """
    if value is None:
        return value, False, False
    raw = str(value).strip()
    if not raw:
        return value, False, False

    match = CODE_RE.match(raw)
    if not match:
        return value, False, True

    name = match.group("name").strip()
    code = match.group("code").strip().upper()
    iso3 = to_iso3(code)
    if not iso3:
        return value, False, True

    new_value = f"{name} ({iso3})"
    return new_value, (new_value != raw), False


def process_file(file_path: Path, apply_changes: bool) -> dict[str, int]:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    stats = {"rows": 0, "changed_cells": 0, "unresolved_cells": 0}

    for row_idx in range(2, ws.max_row + 1):
        stats["rows"] += 1
        # E=进口国, F=出口国
        for col in ("E", "F"):
            cell = ws[f"{col}{row_idx}"]
            new_value, changed, unresolved = normalize_country_cell(cell.value)
            if changed:
                stats["changed_cells"] += 1
                if apply_changes:
                    cell.value = new_value
            if unresolved:
                stats["unresolved_cells"] += 1

    if apply_changes and stats["changed_cells"] > 0:
        wb.save(file_path)
    wb.close()
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert pure_country_pair_usd_count country code to ISO3")
    parser.add_argument("--apply", action="store_true", help="Apply and save changes")
    parser.add_argument("--dry-run", action="store_true", help="Only report, do not save")
    args = parser.parse_args()

    apply_changes = args.apply and not args.dry_run

    if not DATA_DIR.exists():
        raise SystemExit(f"数据目录不存在: {DATA_DIR}")

    files = sorted(DATA_DIR.glob("*.xlsx"))
    if not files:
        raise SystemExit(f"未找到 xlsx: {DATA_DIR}")

    total_rows = 0
    total_changed = 0
    total_unresolved = 0

    print(f"扫描文件: {len(files)} 个")
    for file_path in files:
        stats = process_file(file_path, apply_changes=apply_changes)
        total_rows += stats["rows"]
        total_changed += stats["changed_cells"]
        total_unresolved += stats["unresolved_cells"]
        print(
            f"- {file_path.name}: rows={stats['rows']}, "
            f"changed_cells={stats['changed_cells']}, unresolved_cells={stats['unresolved_cells']}"
        )

    print("\n完成")
    print(f"总行数: {total_rows}")
    print(f"变更单元格: {total_changed}")
    print(f"未识别单元格: {total_unresolved}")
    print(f"模式: {'apply' if apply_changes else 'dry-run'}")


if __name__ == "__main__":
    main()

