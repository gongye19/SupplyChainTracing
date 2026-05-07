from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from .common import (
    add_amount,
    clean_company,
    clean_country,
    empty_amount_bucket,
    normalize_hs,
    parse_decimal,
    parse_int,
    parse_month,
)


EXPECTED_HEADERS = [
    "海关编码",
    "中文产品描述",
    "英文产品描述",
    "月度",
    "进口商",
    "进口商所在国家",
    "出口商",
    "出口商所在国家",
    "原产国",
    "数量",
    "数量单位",
    "金额美元",
    "公吨",
    "交易次数",
]


def iter_trade_files(data_dir: Path) -> list[Path]:
    return [
        path
        for path in sorted(data_dir.glob("*.xlsx"))
        if path.name != "国家对照表.xlsx" and not path.name.startswith("~$")
    ]


def build_aggregates(data_dir: Path) -> tuple[dict[str, int], dict[str, dict]]:
    stats = {
        "rows_seen": 0,
        "rows_accepted": 0,
        "rows_skipped": 0,
        "bad_header_sheets": 0,
        "bad_month_rows": 0,
        "bad_country_rows": 0,
    }

    monthly: dict[tuple, dict[str, Any]] = defaultdict(empty_amount_bucket)
    hs_stats: dict[tuple, dict[str, Any]] = defaultdict(empty_amount_bucket)
    counterparty: dict[tuple, dict[str, Any]] = defaultdict(empty_amount_bucket)
    country_stats: dict[tuple, dict[str, Any]] = defaultdict(empty_amount_bucket)

    files = iter_trade_files(data_dir)
    print(f"找到 {len(files)} 个贸易文件")

    for file_path in files:
        print(f"处理 {file_path.name} ...", flush=True)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
                stats["bad_header_sheets"] += 1
                print(f"  跳过 {ws.title}: 表头不匹配")
                continue

            sheet_accepted = 0
            sheet_skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(value is not None and str(value).strip() for value in row):
                    continue
                stats["rows_seen"] += 1

                raw = list(row[: len(EXPECTED_HEADERS)])
                if len(raw) < len(EXPECTED_HEADERS):
                    raw.extend([None] * (len(EXPECTED_HEADERS) - len(raw)))

                month_pair = parse_month(raw[3])
                importer_country = clean_country(raw[5])
                exporter_country = clean_country(raw[7])
                origin_country = clean_country(raw[8])
                export_side_country = exporter_country or origin_country

                if not month_pair:
                    stats["bad_month_rows"] += 1
                    stats["rows_skipped"] += 1
                    sheet_skipped += 1
                    continue
                if not importer_country or not export_side_country:
                    stats["bad_country_rows"] += 1
                    stats["rows_skipped"] += 1
                    sheet_skipped += 1
                    continue

                year, month = month_pair
                hs_code = normalize_hs(raw[0])
                desc_zh = None if raw[1] is None else str(raw[1]).strip() or None
                desc_en = None if raw[2] is None else str(raw[2]).strip() or None
                importer_name = clean_company(raw[4])
                exporter_name = clean_company(raw[6])
                exporter_company_country = exporter_country or export_side_country
                amount = parse_decimal(raw[11])
                trade_count = parse_int(raw[13])

                add_amount(
                    country_stats[(hs_code, year, month, export_side_country, importer_country)],
                    amount,
                    trade_count,
                    desc_zh,
                    desc_en,
                )

                if importer_name:
                    add_amount(monthly[(importer_name, importer_country, "importer", year, month)], amount, trade_count)
                    add_amount(hs_stats[(importer_name, importer_country, "importer", hs_code)], amount, trade_count, desc_zh, desc_en)
                    if exporter_name:
                        add_amount(
                            counterparty[(importer_name, importer_country, "importer", exporter_name, exporter_company_country)],
                            amount,
                            trade_count,
                        )

                if exporter_name:
                    add_amount(monthly[(exporter_name, exporter_company_country, "exporter", year, month)], amount, trade_count)
                    add_amount(hs_stats[(exporter_name, exporter_company_country, "exporter", hs_code)], amount, trade_count, desc_zh, desc_en)
                    if importer_name:
                        add_amount(
                            counterparty[(exporter_name, exporter_company_country, "exporter", importer_name, importer_country)],
                            amount,
                            trade_count,
                        )

                stats["rows_accepted"] += 1
                sheet_accepted += 1

            print(f"  {ws.title}: 接受 {sheet_accepted:,}，跳过 {sheet_skipped:,}")
        wb.close()

    return stats, {
        "monthly": monthly,
        "hs": hs_stats,
        "counterparty": counterparty,
        "country": country_stats,
    }
